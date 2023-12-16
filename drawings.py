# -*- coding: utf-8 -*-

"""
Module implementing DrawingsMainWindow.
"""

from PyQt6.QtCore import pyqtSlot,Qt		#, pyqtSignal, QEvent
from PyQt6.QtWidgets import QMainWindow, QApplication, QTableWidgetItem, QDialog,QMessageBox,QFileDialog, QMenu #, QAbstractItemView
from Ui_drawings import Ui_DrawingsMainWindow
from Ui_drawingEdit import Ui_Dialog
from PyQt6.QtGui import QColor, QBrush, QIcon
import sys
import os
import pymssql
import openpyxl
from datetime import datetime
import ctypes
import pyperclip
import file_lineEdit
import configparser

#conn=pymssql.connect(host="139.159.180.93",server="HNTESTSERVER",user="mydrawing",password="123456",database="HNdrawings", charset="UTF-8")
myappid="江西晖能新材料图号对比系统——设计：袁伟明"
ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
current_dir=os.path.dirname(os.path.abspath(__file__))  # 获取当前目录
config=configparser.ConfigParser()  # 实例化ConfigParser
config.read(os.path.join(current_dir,"config.ini"))   # 读取当前目录下的config.ini文件
serverhost=config.get("DATABASE","host")    # 读取 [DATABASE] 分组下的 host 的值
servername=config.get("DATABASE","server")
serverdatabase=config.get("DATABASE","database")
serveruser=config.get("DATABASE","username")
serverpassword=config.get("DATABASE","password")
export_path=config.get("PATH","export_path")    # 导出目录
drawing_path=config.get("PATH","drawing_path")        # 图纸目录
wb_config=openpyxl.load_workbook("包装方式.xlsx")
ws_baozhuang = wb_config["包装方式"]
row_baozhuang = tuple(ws_baozhuang.iter_rows(min_row=2, min_col=0, max_col=3))  # 包装方式
wb_config.close()
conn=pymssql.connect(host=serverhost,server=servername,user=serveruser,password=serverpassword,database=serverdatabase, charset="UTF-8")

class DrawingsMainWindow(QMainWindow, Ui_DrawingsMainWindow):
    """
    Class documentation goes here.
    """
    #clicked=pyqtSignal()
    def __init__(self, parent=None):
        super(DrawingsMainWindow, self).__init__(parent)
        self.setupUi(self)
        self.setWindowIcon(QIcon("drawing.ico"))
        self.tableWidget.verticalHeader().sectionClicked.connect(self.VerSectionClicked)#表头行单击信号
        self.tableWidget.horizontalHeader().sectionClicked.connect(self.HorSectionClicked)#表头列单击信号
        # table widget 右键菜单 放在主窗口__init__(self):下
        self.tableWidget.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu) # 允许右键产生子菜单
        self.tableWidget.customContextMenuRequested.connect(self.tableWidget_menu)  # 右键菜单

    def tableWidget_menu(self, pos):    # 右键菜单
        """
        :return:
        """
        row_num = -1
        for i in self.tableWidget.selectionModel().selection().indexes():
            row_num = i.row()

        if row_num >=0: # 表格生效的行数，小于0行点击右键，不会弹出菜单
            menu = QMenu() #实例化菜单
            item1 = menu.addAction(u"复制当前料号")
            menu.addSeparator()
            item2 = menu.addAction(u"修改选定数据")
            item3 = menu.addAction(u"删除选定记录")
            menu.addSeparator()
            item4=menu.addAction(u"打开产品图纸")
            item5=menu.addAction(u"打开生产图纸")
            item6=menu.addAction(u"打开客户图纸")
            action = menu.exec(self.tableWidget.mapToGlobal(pos))
        else:
            return
        if action == item1:
            #print(self.tableWidget.item(row_num, 1).text())
            pyperclip.copy(self.tableWidget.item(row_num, 1).text())
        elif action == item2:
            self.on_modifyButton_clicked()
        elif action == item3:
            self.on_deleteButton_clicked()
        elif action==item4:
            if self.tableWidget.item(row_num, 10).text()!="":
                if self.tableWidget.item(row_num, 10).text()[1:3]!=":/":
                    os.startfile(f"{drawing_path}{self.tableWidget.item(row_num, 10).text()}")
                else:
                    os.startfile(f"{self.tableWidget.item(row_num, 10).text()}")
            else:
                pass
        elif action==item5:
            if self.tableWidget.item(row_num, 11).text()!="":
                if self.tableWidget.item(row_num, 11).text()[1:3]!=":/":
                    os.startfile(f"{drawing_path}{self.tableWidget.item(row_num, 11).text()}")
                else:
                    os.startfile(f"{self.tableWidget.item(row_num, 11).text()}")
            else:
               pass
        elif action==item6:
            if self.tableWidget.item(row_num, 12).text()!="":
                if self.tableWidget.item(row_num, 12).text()[1:3]!=":/":
                    os.startfile(f"{drawing_path}{self.tableWidget.item(row_num, 12).text()}")
                else:
                    os.startfile(f"{self.tableWidget.item(row_num, 12).text()}")
            else:
               pass

    def VerSectionClicked(self,index):
        #print (index)
        pass
    def HorSectionClicked(self,index):
        #print (index)
        if self.tableWidget.rowCount()>0:
            self.tableWidget.sortItems(index, Qt.SortOrder.AscendingOrder)  # 升序排列
        
        
    
    def choose_combo(self,xinghao):
        baozhuang_list = []
        if "V8" in xinghao or "保温箱" in xinghao:
            for bz in row_baozhuang:
                if "V8" in bz[2].value:
                    baozhuang_list.append(bz[1].value)
        elif "M5600" in xinghao:
            for bz in row_baozhuang:
                if "M5600" in bz[2].value:
                    baozhuang_list.append(bz[1].value)
        elif "M59" in xinghao:
            #if "M5902" in xinghao:
                #self.peifang_comboBox.setCurrentIndex(1)
            #else:
                #self.peifang_comboBox.setCurrentIndex(0)
            for bz in row_baozhuang:
                if "M59" in bz[2].value:
                    baozhuang_list.append(bz[1].value)
        elif "M57" in xinghao:
            for bz in row_baozhuang:
                if "M57" in bz[2].value:
                    baozhuang_list.append(bz[1].value)
        elif "M52" in xinghao:
            for bz in row_baozhuang:
                if "M52" in bz[2].value:
                    baozhuang_list.append(bz[1].value)
        elif "M51" in xinghao:
            for bz in row_baozhuang:
                if "M51" in bz[2].value:
                    baozhuang_list.append(bz[1].value)
        elif "M3100" in xinghao:
            for bz in row_baozhuang:
                if "M3100" in bz[2].value:
                    baozhuang_list.append(bz[1].value)
        else:
            baozhuang_list.append("")
        self.comboBox_2.clear()
        self.comboBox_2.addItems(baozhuang_list)       # 选择主分类后重写包装方式列表框
        #self.comboBox_2.setCurrentIndex(0)
        #print("lineEdit_4丢失焦点")

    @pyqtSlot()
    def on_findButton_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        tj=" 1=1"
        if self.PartNumber_lineEdit.text()!="":
            tj=" PartNumber like '%"+self.PartNumber_lineEdit.text()+"%'"
        if self.ProductName_lineEdit.text()!="":
            if tj==" 1=1":
                tj=" ProductName like '%"+self.ProductName_lineEdit.text()+"%'"
            else:
                tj=tj+" and "+" ProductName like '%"+self.ProductName_lineEdit.text()+"%'"
        if self.Specs_lineEdit.text()!="":
            if tj==" 1=1":
                tj=" Specs like '%"+self.Specs_lineEdit.text()+"%'"
            else:
                tj=tj+" and "+" Specs like '%"+self.Specs_lineEdit.text()+"%'"
        if self.TemperatureLevel_lineEdit.text()!="":
            if tj==" 1=1":
                tj=" TemperatureLevel like '%"+self.TemperatureLevel_lineEdit.text()+"%'"
            else:
                tj=tj+" and "+" TemperatureLevel like '%"+self.TemperatureLevel_lineEdit.text()+"%'"
        if self.Density_lineEdit.text()!="":
            if tj==" 1=1":
                tj=" Density like '%"+self.Density_lineEdit.text()+"%'"
            else:
                tj=tj+" and "+" Density like '%"+self.Density_lineEdit.text()+"%'"
        if self.PackagingMethod_lineEdit.text()!="":
            if tj==" 1=1":
                tj=" PackagingMethod like '%"+self.PackagingMethod_lineEdit.text()+"%'"
            else:
                tj=tj+" and "+" PackagingMethod like '%"+self.PackagingMethod_lineEdit.text()+"%'"
        if self.Drawing_N_lineEdit.text()!="":
            if tj==" 1=1":
                tj=" Drawing_N like '%"+self.Drawing_N_lineEdit.text()+"%'"
            else:
                tj=tj+" and "+" Drawing_N like '%"+self.Drawing_N_lineEdit.text()+"%'"
        if self.Drawing_W_lineEdit.text()!="":
            if tj==" 1=1":
                tj=" Drawing_W like '%"+self.Drawing_W_lineEdit.text()+"%'"
            else:
                tj=tj+" and "+" Drawing_W like '%"+self.Drawing_W_lineEdit.text()+"%'"
        if self.Description_lineEdit.text()!="":
            if tj==" 1=1":
                tj=" Description like '%"+self.Description_lineEdit.text()+"%'"
            else:
                tj=tj+" and "+" Description like '%"+self.Description_lineEdit.text()+"%'"
        if self.ProductDrawings_lineEdit.text()!="":
            if tj==" 1=1":
                tj=" ProductDrawings like '%"+self.ProductDrawings_lineEdit.text()+"%'"
            else:
                tj=tj+" and "+" ProductDrawings like '%"+self.ProductDrawings_lineEdit.text()+"%'"
        if self.DesignDrawings_lineEdit.text()!="":
            if tj==" 1=1":
                tj=" DesignDrawings like '%"+self.DesignDrawings_lineEdit.text()+"%'"
            else:
                tj=tj+" and "+" DesignDrawings like '%"+self.DesignDrawings_lineEdit.text()+"%'"
        if self.CustomerDrawings_lineEdit.text()!="":
            if tj==" 1=1":
                tj=" CustomerDrawings like '%"+self.CustomerDrawings_lineEdit.text()+"%'"
            else:
                tj=tj+" and "+" CustomerDrawings like '%"+self.CustomerDrawings_lineEdit.text()+"%'"

        sql="select ID,PartNumber,ProductName,Specs,TemperatureLevel,Density,PackagingMethod,Drawing_N,Drawing_W,Description,ProductDrawings,DesignDrawings,CustomerDrawings,Remark from DrawingsData where"+tj+"ORDER BY PartNumber"   # 组合查询条件
        
        if conn:
            cursor=conn.cursor()
            cursor.execute(sql)
            #self.tableWidget.clearContents()   # 清空tablewidger中的内容，不包括表头
            self.tableWidget.setRowCount(0) #设置tableWidgetb行数为0
            for row in cursor:
                self.tableWidget.setRowCount(cursor.rowcount+1) # 设置行数
                for i in range(14):
                    self.tableWidget.setItem(cursor.rowcount, i, QTableWidgetItem(str(row[i]) if row[i]!=None else "")) # 插入一行记录  
                    if cursor.rowcount%2==1:    # 隔行设置颜色
                        self.tableWidget.item(cursor.rowcount, i).setBackground(QBrush(QColor(240,240,240)))
            for table_width in range(10):
                self.tableWidget.resizeColumnToContents(table_width)  #设置自动列宽

            self.tableWidget.horizontalHeader().setStretchLastSection(True) # 设置最后一列自动填充容器
            # QTableWidget设置整行选中
            #self.tableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)
            #self.tableWidget.setSelectionMode(QAbstractItemView.SingleSelection)
            self.statusBar.showMessage("查询到"+str(self.tableWidget.rowCount())+"条记录...") #在状态栏显示查询到的记录数

    @pyqtSlot()
    def on_addButton_clicked(self):
        """
        Slot documentation goes here.
        """
        drawingEdit=show_dialog()       # 加载Dialog窗体
        self.addButton.clicked.connect(drawingEdit.drawingEditOpen)     # 信号槽连接并打开Dialog窗体
        drawingEdit.lineEdit_2.setFocus()
        
        file_lineEdit.lineEdit_dragFile_injector(drawingEdit.lineEdit_11)
        file_lineEdit.lineEdit_dragFile_injector(drawingEdit.lineEdit_12)
        file_lineEdit.lineEdit_dragFile_injector(drawingEdit.lineEdit_13)
        if drawingEdit.exec():      # 如果窗体是点击OK退出
            self.data=(drawingEdit.lineEdit_2.text(),drawingEdit.lineEdit_3.text(),drawingEdit.lineEdit_4.text(),drawingEdit.comboBox.currentText(),drawingEdit.lineEdit_6.text(),drawingEdit.comboBox_2.currentText(),drawingEdit.lineEdit_8.text(),drawingEdit.lineEdit_9.text(),drawingEdit.lineEdit_10.text(),drawingEdit.lineEdit_11.text(),drawingEdit.lineEdit_12.text(),drawingEdit.lineEdit_13.text(),drawingEdit.lineEdit_14.text())
            insert_sql="INSERT INTO DrawingsData (PartNumber,ProductName,Specs,TemperatureLevel,Density,PackagingMethod,Drawing_N,Drawing_W,Description,ProductDrawings,DesignDrawings,CustomerDrawings,Remark) VALUES "+str(self.data)
            #print(insert_sql)
            insert_cursor=conn.cursor() # 使用cursor()方法获取操作游标
            try:
                insert_cursor.execute(insert_sql)   # 提交sql语句
                #print (insert_sql)
                conn.commit()   # 执行sql语句
                self.statusBar.showMessage("成功插入1条记录...请重新查询显示新增数据...")
                insert_cursor.close()    
            except:
                conn.rollback()     # 发生错误时回滚            
        

    @pyqtSlot()
    def on_deleteButton_clicked(self):
        try:
            row_lst=[]
            # 获取选中的行号并添加到列表row_lst
            for item in self.tableWidget.selectedItems():
                cur_row=item.row()
                if cur_row in row_lst:
                    continue
                else:
                    row_lst.append(cur_row)
            #print(row_lst)
            id_lst=[]
            if len(row_lst)>0:
                for x in row_lst:
                    id_lst.append(int(self.tableWidget.item(x, 0).text()))
                    
                reply=QMessageBox.question(self,"确认提示！", "确认要删除选中的ID为%s的数据吗？" % sorted(id_lst))
                reply=str(reply)
                #print(id_lst)
                if reply=="StandardButton.Yes":
                    #print("确认删除")
                    try:
                        for i in range(len(id_lst)):
                            del_sql="DELETE from DrawingsData WhERE ID=%s" % id_lst[i]
                            print(del_sql)
                            del_cursor=conn.cursor()
                            del_cursor.execute(del_sql)
                            conn.commit() 
                            self.statusBar.showMessage("删除ID为"+str(id_lst[i])+"的记录......")
                            del_cursor.close()
                        self.statusBar.showMessage("删除ID为"+str(sorted(id_lst))+"的记录......")
                    except:
                        self.statusBar.showMessage("删除出现错误...请检查...")
        except:
            QMessageBox.information(self, "提示", "操作出现错误")
            

    @pyqtSlot()
    def on_addAllButton_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        file_path=QFileDialog.getOpenFileName(self, "选择drawings模板文件", ".", "文件类型(*.xlsx *.xls)")
        #print(file_path)
        if len(file_path[0])>1:
            wb=openpyxl.load_workbook(file_path[0])
            ws=wb.active
            rows=ws.max_row     #总行数
            #columns=ws.max_column     #总列数

            if rows>0:
                add_sql="INSERT INTO DrawingsData (PartNumber,ProductName,Specs,TemperatureLevel,Density,PackagingMethod,Drawing_N,Drawing_W,Description,ProductDrawings,DesignDrawings,CustomerDrawings,Remark) VALUES "
                for i in range(rows-1):
                    temp_list=[]
                    for each in ws.iter_cols(min_row=2):
                        if each[i].value==None:
                            each[i].value=""
                        temp_list.append(each[i].value)
                    row_sql=add_sql+str(tuple(temp_list))
                    try:
                        add_cursor=conn.cursor()
                        add_cursor.execute(row_sql)
                        conn.commit()
                        add_cursor.close()
                        self.statusBar.showMessage("正在批量新增记录......当前成功插入第%d条记录...请等待操作完成..." % (i+1))
                    except:
                        conn.rollback()
                    #print(add_sql+str(tuple(temp_list)))
                self.statusBar.showMessage("成功插入%d条记录...请重新查询显示新增数据..." % (rows-1))
            wb.close()
        else:
            #QMessageBox.information(self, "提示", "未选择模板文件......")
            self.statusBar.showMessage("未选择模板文件.........")

    @pyqtSlot()
    def on_deleteAllButton_clicked(self):       # 批量修改记录
        """
        Slot documentation goes here.
        """
        file_path=QFileDialog.getOpenFileName(self, "选择drawings模板文件", "./导出文件", "文件类型(*.xlsx *.xls)")
        #print(file_path)
        if len(file_path[0])>1:
            wb=openpyxl.load_workbook(file_path[0])
            ws=wb.active
            rows=ws.max_row     #总行数
            #columns=ws.max_column     #总列数

            if rows>1:
                #add_sql="UPDATE DrawingsData SET PartNumber='"+modify_data[1]+"',ProductName='"+modify_data[2]+"',Specs='"+modify_data[3]+"',Drawing_N='"+modify_data[4]+"',Drawing_W='"+modify_data[5]+"',Remark='"+modify_data[6]+"' WHERE ID=%d"
                for i in range(rows-1):
                    temp_list=[]
                    for each in ws.iter_cols(min_row=2):
                        if each[i].value==None:
                            each[i].value=""
                        temp_list.append(each[i].value)
                    #print(temp_list)
                    row_sql="UPDATE DrawingsData SET PartNumber='"+temp_list[1]+"',ProductName='"+temp_list[2]+"',Specs='"+temp_list[3]+"',TemperatureLevel='"+temp_list[4]+"',Density='"+temp_list[5]+"',PackagingMethod='"+temp_list[6]+"',Drawing_N='"+temp_list[7]+"',Drawing_W='"+temp_list[8]+"',Description='"+temp_list[9]+"',ProductDrawings='"+temp_list[10]+"',DesignDrawings='"+temp_list[11]+"',CustomerDrawings='"+temp_list[12]+"',Remark='"+temp_list[13]+"' WHERE ID=%d" % int(temp_list[0])
                    #print(row_sql)
                    
                    try:
                        add_cursor=conn.cursor()
                        add_cursor.execute(row_sql)
                        conn.commit()
                        add_cursor.close()
                        self.statusBar.showMessage("正在批量修改记录......当前成功修改第%d条记录...请等待操作完成..." % (i+1))
                    except:
                        conn.rollback()
                    #print(add_sql+str(tuple(temp_list)))
                self.statusBar.showMessage("成功修改%d条记录...请重新查询显示新增数据..." % (rows-1))
                
            wb.close()
        else:
            #QMessageBox.information(self, "提示", "未选择模板文件......")
            self.statusBar.showMessage("未选择模板文件.........")

    @pyqtSlot()
    def on_modifyButton_clicked(self):
        """
        Slot documentation goes here.
        """
        #clicked=pyqtSignal()
        #print(self.tableWidget.currentRow())    # 获取当前光标所在行，从0开始计数
        if len(self.tableWidget.selectedItems())!=0:
            row_num=self.tableWidget.currentRow()
            #print(row_num)
            self.drawingModify=show_dialog()       # 加载Dialog窗体
            #self.modifyButton.clicked.connect(drawingModify.drawingEditOpen)     # 信号槽连接并打开Dialog窗体
            self.drawingModify.drawingEditOpen()

            id_old=int(self.tableWidget.item(row_num,0).text())
            self.drawingModify.lineEdit.setText(self.tableWidget.item(row_num,0).text())
            self.drawingModify.lineEdit_2.setText(self.tableWidget.item(row_num,1).text())
            self.drawingModify.lineEdit_3.setText(self.tableWidget.item(row_num,2).text())
            self.drawingModify.lineEdit_4.setText(self.tableWidget.item(row_num,3).text())
            self.drawingModify.comboBox.setCurrentText(self.tableWidget.item(row_num,4).text())
            self.drawingModify.lineEdit_6.setText(self.tableWidget.item(row_num,5).text())
            self.drawingModify.comboBox_2.setCurrentText(self.tableWidget.item(row_num,6).text())
            self.drawingModify.lineEdit_8.setText(self.tableWidget.item(row_num,7).text())
            self.drawingModify.lineEdit_9.setText(self.tableWidget.item(row_num,8).text())
            self.drawingModify.lineEdit_10.setText(self.tableWidget.item(row_num, 9).text())
            self.drawingModify.lineEdit_11.setText(self.tableWidget.item(row_num, 10).text())
            self.drawingModify.lineEdit_12.setText(self.tableWidget.item(row_num, 11).text())
            self.drawingModify.lineEdit_13.setText(self.tableWidget.item(row_num, 12).text())
            self.drawingModify.lineEdit_14.setText(self.tableWidget.item(row_num, 13).text())
            file_lineEdit.lineEdit_dragFile_injector(self.drawingModify.lineEdit_11)
            file_lineEdit.lineEdit_dragFile_injector(self.drawingModify.lineEdit_12)
            file_lineEdit.lineEdit_dragFile_injector(self.drawingModify.lineEdit_13)
            #self.drawingModify.lineEdit_4.clicked.connect(self.choose_combo)
            if self.drawingModify.exec():      # 如果窗体是点击OK退出
                try:
                    modify_data=(self.drawingModify.lineEdit.text(),self.drawingModify.lineEdit_2.text(),self.drawingModify.lineEdit_3.text(),self.drawingModify.lineEdit_4.text(),self.drawingModify.comboBox.currentText(),self.drawingModify.lineEdit_6.text(),self.drawingModify.comboBox_2.currentText(),self.drawingModify.lineEdit_8.text(), self.drawingModify.lineEdit_9.text(), self.drawingModify.lineEdit_10.text(), self.drawingModify.lineEdit_11.text(), self.drawingModify.lineEdit_12.text(), self.drawingModify.lineEdit_13.text(), self.drawingModify.lineEdit_14.text())
                    modify_sql="UPDATE DrawingsData SET PartNumber='"+modify_data[1]+"',ProductName='"+modify_data[2]+"',Specs='"+modify_data[3]+"',TemperatureLevel='"+modify_data[4]+"',Density='"+modify_data[5]+"',PackagingMethod='"+modify_data[6]+"',Drawing_N='"+modify_data[7]+"',Drawing_W='"+modify_data[8]+"',Description='"+modify_data[9]+"',ProductDrawings='"+modify_data[10]+"',DesignDrawings='"+modify_data[11]+"',CustomerDrawings='"+modify_data[12]+"',Remark='"+modify_data[13]+"' WHERE ID=%d" % id_old
                    #print(modify_sql, id_old)
                    modify_cursor=conn.cursor()
                    modify_cursor.execute(modify_sql)
                    conn.commit()
                    self.statusBar.showMessage("成功修改1条记录...请重新查询显示数据...")
                    modify_cursor.close()
                except:
                    conn.rollback()     # 发生错误时回滚
                    QMessageBox.warning(self, "提示", "数据修改错误，请重试......")
        else:
            QMessageBox.information(self, "提示", "未选择记录......")
        #for x in self.tableWidget.selectedItems():
            #print(x.text())

    @pyqtSlot()
    def on_exitButton_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        conn.close()
        self.exitButton.clicked.connect(sys.exit(app.exec()))

    @pyqtSlot()
    def on_exportButton_clicked(self):
        now=datetime.now()
        nowtime=str(now.date()).replace("-","")+str(now.strftime("%H%M%S"))
        excelfilename=export_path+"图纸图号导出"+nowtime+".xlsx"
        filename="图纸图号导出"+nowtime+".xlsx"
        wb=openpyxl.Workbook()
        ws=wb.create_sheet("图号对比", 0)
        #print(ws)
        new_header=["ID", "料号", "品名", "规格", "温度等级", "密度", "包装方式", "内部图号", "客户图号", "物料描述", "产品图纸", "生产图纸", "客户图纸", "备注"]
        ws.append(new_header)   # 写入一行数据
        column_width=[6, 12, 20, 30,10,6,20, 35,40,40, 40,40, 40,  20]
        for i in range(1, 14):   # 设置所有列宽
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width[i-1]
            #print(ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width)
        num=self.tableWidget.rowCount() # 获取qtableWidget总行数
        for n in range(num):
            temp_data=[]
            for x in range(14):
                temp_data.append(self.tableWidget.item(n,x).text())
            #print(temp_data)
            ws.append(temp_data)
            
        reply=QMessageBox.question(self,"确认提示！", "确认要导出查询结果为《%s》的Excel文件吗？" % filename)
        reply=str(reply)
        #print(id_lst)
        if reply=="StandardButton.Yes":
            wb.save(excelfilename)
            wb.close()
            self.statusBar.showMessage("%s 导出成功......" % filename) 
        else:
            wb.close()
            self.statusBar.showMessage("取消导出文件......") 
''''
class QLineEditDropHandler(QObject):    
    def eventFilter(self, watched, event):
        if event.type() == QEvent.DragEnter:
            # 我们需要明确接受此事件才能接收 QDropEvents！
            event.accept()
        if event.type() == QEvent.Drop:
            md = event.mimeData()
            if md.hasUrls():
                url=md.urls()[0]
                watched.setText(url.toLocalFile())
                return True
        return super().eventFilter(watched, event)
 ''' 
class show_dialog(QDialog, Ui_Dialog):
    #clicked=pyqtSignal()
    def __init__(self):
        super(show_dialog, self).__init__()
        self.setupUi(self)
        #self.data=()
        self.setWindowIcon(QIcon("drawing.ico"))    # 设置图标
        #xinhao=self.lineEdit_3.text()
        self.setWindowFlag(Qt.WindowType.WindowStaysOnTopHint, True)    # 设置窗口始终显示在最前
        self.activateWindow()   # 激活窗口
        self.lineEdit_4.textChanged.connect(lambda:DrawingsMainWindow.choose_combo(self, xinghao=self.lineEdit_3.text()))
        self.lineEdit_8.setAcceptDrops(True)
        self.lineEdit_8.setDragEnabled(True) 
        #self.lineEdit_4.clicked.connect(DrawingsMainWindow.choose_combo)
        #self.lineEdit_8.installEventFilter(QEventHandler(self))
        #self.lineEdit_9.installEventFilter(QEventHandler(self))
    """
    # 鼠标拖入事件
    def dragEnterEvent(self, event):
        #self.setWindowTitle('dragEnterEvent')
        file = event.mimeData().urls()[0].toLocalFile()  # ==> 获取文件路径
        #print(file)
        if file not in self.paths:  # ==> 去重显示
            print("拖拽的文件 ==> {}".format(file))
            self.paths += file + "\n"
            self.lineEdit_8.setText(self.paths)
            # 鼠标放开函数事件
            event.accept()
    """
    #def eventFilter(self, widget, event):
        #if widget==self.edit:
            #if event.type()==QEvent.FocusOut:
                #pass
            #elif event.type()==QEvent.FocusIn:
                #self.clicked.emitemit()
            #else:
                #pass
        #return False

    def drawingEditOpen(self):
        self.show()
        #self.data=(drawingEdit.lineEdit.text(),drawingEdit.lineEdit_2.text(),drawingEdit.lineEdit_3.text(),drawingEdit.lineEdit_4.text(),drawingEdit.lineEdit_5.text(),drawingEdit.lineEdit_6.text(),drawingEdit.lineEdit_7.text())
        
        
#class mylineedit(QLineEdit):
	#clicked=pyqtSignal()	# 定义clicked信号
	#def mouseReleaseEvent(self, QMouseEvent):
		#if QMouseEvent.button()==Qt.LeftButton:
			#self.clicked.emit()	# 发送点击信号

if __name__ == "__main__":
    app = QApplication(sys.argv)
    drawingsWindow = DrawingsMainWindow()
    drawingsWindow.show()
    #drawingsWindow.setWindowFlags(Qt.WindowType.WindowMaximizeButtonHint|Qt.WindowType.MSWindowsFixedSizeDialogHint)
    #drawingsWindow.setFixedSize(drawingsWindow.width(), drawingsWindow.height())
    sys.exit(app.exec())
